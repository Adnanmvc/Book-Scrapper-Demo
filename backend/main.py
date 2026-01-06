from fastapi import FastAPI, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

from pydantic import BaseModel
from io import BytesIO
from openpyxl import Workbook

from scraper.books import crawl_all_pages, Book

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], 
    allow_credentials=True,
    allow_methods=["*"], 
    allow_headers=["*"],
)


class BookOut(BaseModel):
    title: str
    price: float
    availability: str
    rating: int
    url: str


@app.get("/api/books")
def get_books_xlsx(
    limit: int = Query(50, ge=1, le=1000),
):
    books: list[BookOut] = []
    for book in crawl_all_pages():
        books.append(BookOut(**book.__dict__))
        if len(books) >= limit:
            break

    if not books:
        wb = Workbook()
        ws = wb.active
        ws.title = "Books"
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="books.xlsx"'},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    headers = ["title", "price", "availability", "rating", "url"]
    ws.append(headers)

    for b in books:
        ws.append([
            b.title,
            b.price,
            b.availability,
            b.rating,
            b.url,
        ])

    for cell in ws["B"][1:]:
        cell.number_format = "#,##0.00"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="books.xlsx"'
        },
    )

@app.get("/api/book", response_model=BookOut)
def get_book(
    title: str
):
    for book in crawl_all_pages():
        if book.title == title:
            return BookOut(
                title=book.title,
                price=book.price,
                availability=book.availability,
                rating=book.rating,
                url=book.url
            )
    
    raise HTTPException(status_code=404, detail=f"Book with title '{title}' not found")
        